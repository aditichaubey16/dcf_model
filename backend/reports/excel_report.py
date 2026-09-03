"""Excel DCF workbook generator — historicals, cash-flow change analysis,
DCF assumptions/projection, and a WACC x terminal-growth sensitivity grid,
laid out the way a PE deal-team model is typically tabbed.
"""
from __future__ import annotations

import io
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

NAVY = "1F2A44"
LIGHT_GREY = "F2F4F7"
ACCENT = "2A78D6"

HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
TITLE_FONT = Font(bold=True, size=16, color=NAVY)
SUBTITLE_FONT = Font(italic=True, size=10, color="666666")
LABEL_FONT = Font(bold=True)
THIN = Side(style="thin", color="D0D5DD")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _header_row(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        cell = ws.cell(row=row, column=start_col + i, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER


def _autofit(ws, min_width=10, max_width=42):
    for col_cells in ws.columns:
        length = max((len(str(c.value)) for c in col_cells if c.value is not None), default=0)
        col = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col].width = min(max(length + 2, min_width), max_width)


def build_excel_report(
    *,
    company_name: str,
    periods: list[str],
    metrics: list[dict],
    changes: list[dict],
    change_flags: list[dict],
    concerns: list[dict],
    dcf: dict,
    assumptions: dict,
    scenarios: dict | None = None,
) -> bytes:
    wb = Workbook()

    # ---------------- Summary ----------------
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = f"{company_name} — DCF Valuation Summary"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"Generated {date.today().isoformat()} · Confidential — prepared for internal review"
    ws["A2"].font = SUBTITLE_FONT

    summary_rows = [
        ("Enterprise Value", dcf.get("enterprise_value")),
        ("Less: Net Debt", -dcf.get("net_debt", 0) if dcf.get("net_debt") else 0),
        ("Equity Value", dcf.get("equity_value")),
        ("Value per Share", dcf.get("value_per_share")),
        ("", ""),
        ("WACC", assumptions.get("wacc")),
        ("Terminal Growth Rate", assumptions.get("terminal_growth")),
        ("Revenue Growth Assumption", assumptions.get("revenue_growth_rate")),
        ("Assumed FCF Margin", dcf.get("assumed_fcf_margin")),
        ("Historical Avg FCF Margin", dcf.get("historical_fcf_margin")),
        ("Projection Years", assumptions.get("projection_years")),
    ]
    r = 4
    for label, val in summary_rows:
        ws.cell(row=r, column=1, value=label).font = LABEL_FONT
        ws.cell(row=r, column=2, value=val)
        r += 1

    if dcf.get("warnings"):
        r += 1
        ws.cell(row=r, column=1, value="Model warnings:").font = LABEL_FONT
        for w in dcf["warnings"]:
            r += 1
            ws.cell(row=r, column=1, value=f"- {w}")

    _autofit(ws)

    # ---------------- Historicals ----------------
    ws2 = wb.create_sheet("Historicals & Ratios")
    fields = [
        "period", "revenue", "revenue_growth", "net_income", "net_income_growth",
        "gross_margin", "operating_margin", "net_margin",
        "operating_cash_flow", "investing_cash_flow", "financing_cash_flow", "free_cash_flow",
        "current_ratio", "quick_ratio", "debt_to_equity", "interest_coverage", "roe", "roa",
        "total_assets", "total_liabilities", "total_equity",
    ]
    _header_row(ws2, 1, [f.replace("_", " ").title() for f in fields])
    for i, m in enumerate(metrics, start=2):
        for j, f in enumerate(fields, start=1):
            cell = ws2.cell(row=i, column=j, value=m.get(f))
            cell.border = BORDER
    _autofit(ws2)
    ws2.freeze_panes = "A2"

    # ---------------- Cash Flow Change Analysis ----------------
    ws3 = wb.create_sheet("Cash Flow Changes")
    cf_fields = [
        "period", "operating_cf", "operating_cf_change_pct", "investing_cf", "investing_cf_change_pct",
        "financing_cf", "financing_cf_change_pct", "free_cash_flow", "free_cash_flow_change_pct",
        "revenue_growth_pct", "fcf_vs_revenue_divergence", "fcf_conversion", "ocf_to_ni",
    ]
    _header_row(ws3, 1, [f.replace("_", " ").title() for f in cf_fields])
    for i, c in enumerate(changes, start=2):
        for j, f in enumerate(cf_fields, start=1):
            cell = ws3.cell(row=i, column=j, value=c.get(f))
            cell.border = BORDER
    _autofit(ws3)
    ws3.freeze_panes = "A2"

    start = len(changes) + 3
    ws3.cell(row=start, column=1, value="Flagged changes").font = LABEL_FONT
    _header_row(ws3, start + 1, ["Severity", "Period", "Title", "Message"])
    for i, fl in enumerate(change_flags, start=start + 2):
        ws3.cell(row=i, column=1, value=fl.get("severity"))
        ws3.cell(row=i, column=2, value=fl.get("period"))
        ws3.cell(row=i, column=3, value=fl.get("title"))
        ws3.cell(row=i, column=4, value=fl.get("message"))
    _autofit(ws3)

    # ---------------- DCF Projection ----------------
    ws4 = wb.create_sheet("DCF Projection")
    ws4["A1"] = "Free Cash Flow Projection"
    ws4["A1"].font = LABEL_FONT
    _header_row(ws4, 2, ["Year", "Projected Revenue", "Projected FCF", "Discount Factor", "Present Value"])
    for i, p in enumerate(dcf.get("projections", []), start=3):
        ws4.cell(row=i, column=1, value=p.get("year"))
        ws4.cell(row=i, column=2, value=p.get("revenue"))
        ws4.cell(row=i, column=3, value=p.get("free_cash_flow"))
        ws4.cell(row=i, column=4, value=p.get("discount_factor"))
        ws4.cell(row=i, column=5, value=p.get("present_value"))
    end = 3 + len(dcf.get("projections", []))
    ws4.cell(row=end + 1, column=1, value="Sum of PV of FCF").font = LABEL_FONT
    ws4.cell(row=end + 1, column=5, value=dcf.get("sum_pv_fcf"))
    ws4.cell(row=end + 2, column=1, value="Terminal Value").font = LABEL_FONT
    ws4.cell(row=end + 2, column=5, value=dcf.get("terminal_value"))
    ws4.cell(row=end + 3, column=1, value="PV of Terminal Value").font = LABEL_FONT
    ws4.cell(row=end + 3, column=5, value=dcf.get("pv_terminal_value"))
    ws4.cell(row=end + 4, column=1, value="Enterprise Value").font = LABEL_FONT
    ws4.cell(row=end + 4, column=5, value=dcf.get("enterprise_value"))
    _autofit(ws4)

    # ---------------- Sensitivity ----------------
    ws5 = wb.create_sheet("Sensitivity (EV)")
    sens = dcf.get("sensitivity", {})
    wacc_axis = sens.get("wacc_axis", [])
    g_axis = sens.get("terminal_growth_axis", [])
    ev_grid = sens.get("enterprise_values", [])

    ws5.cell(row=1, column=1, value="WACC \\ Terminal Growth").font = LABEL_FONT
    for j, g in enumerate(g_axis, start=2):
        cell = ws5.cell(row=1, column=j, value=g)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    for i, w in enumerate(wacc_axis, start=2):
        cell = ws5.cell(row=i, column=1, value=w)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        for j, val in enumerate(ev_grid[i - 2] if i - 2 < len(ev_grid) else [], start=2):
            c = ws5.cell(row=i, column=j, value=val)
            c.border = BORDER
    _autofit(ws5)

    # ---------------- Scenario Analysis ----------------
    if scenarios:
        ws_scen = wb.create_sheet("Scenario Analysis")
        _header_row(ws_scen, 1, ["Metric", "Bear", "Base", "Bull"])
        rows_def = [
            ("Revenue Growth Rate", "revenue_growth_rate"),
            ("Assumed FCF Margin", "assumed_fcf_margin"),
            ("Enterprise Value", "enterprise_value"),
            ("Equity Value", "equity_value"),
            ("Value per Share", "value_per_share"),
        ]
        order = ["bear", "base", "bull"]
        for i, (label, field_name) in enumerate(rows_def, start=2):
            ws_scen.cell(row=i, column=1, value=label).font = LABEL_FONT
            for j, case in enumerate(order, start=2):
                ws_scen.cell(row=i, column=j, value=scenarios.get(case, {}).get(field_name))
        _autofit(ws_scen)

    # ---------------- Concerns ----------------
    ws6 = wb.create_sheet("Concern Areas")
    _header_row(ws6, 1, ["Severity", "Period", "Title", "Message"])
    for i, c in enumerate(concerns, start=2):
        ws6.cell(row=i, column=1, value=c.get("severity"))
        ws6.cell(row=i, column=2, value=c.get("period"))
        ws6.cell(row=i, column=3, value=c.get("title"))
        ws6.cell(row=i, column=4, value=c.get("message"))
    _autofit(ws6)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
